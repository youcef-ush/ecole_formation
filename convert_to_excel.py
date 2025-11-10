"""
Script de conversion PROJECT_PLAN.csv vers PROJECT_PLAN.xlsx
Avec formatage professionnel Excel
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Lire le CSV
print("📖 Lecture du fichier CSV...")
df = pd.read_csv('PROJECT_PLAN.csv', encoding='utf-8')

# 2. Créer le fichier Excel
print("💾 Création du fichier Excel...")
df.to_excel('PROJECT_PLAN.xlsx', index=False, sheet_name='Plan Projet')

# 3. Charger le workbook pour le formatage
print("🎨 Formatage du fichier Excel...")
wb = load_workbook('PROJECT_PLAN.xlsx')
ws = wb['Plan Projet']

# Couleurs
HEADER_COLOR = "366092"  # Bleu foncé
HAUTE_COLOR = "FF6B6B"   # Rouge pour priorité HAUTE
MOYENNE_COLOR = "FFA500"  # Orange pour priorité MOYENNE
BASSE_COLOR = "4ECDC4"   # Vert pour priorité BASSE
PHASE_COLORS = {
    "1. Base de Données": "E3F2FD",
    "2. Services Backend": "F3E5F5",
    "3. API Routes": "E8F5E9",
    "4. Frontend - Scanner": "FFF3E0",
    "5. Frontend - Gestion": "FCE4EC",
    "6. Dashboard Stats": "F1F8E9",
    "7. Tests & Validation": "FFEBEE",
    "8. Documentation": "E0F2F1"
}

# Style de bordure
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# 4. Formater l'en-tête
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# 5. Ajuster largeur des colonnes
column_widths = {
    'A': 8,   # ID
    'B': 22,  # Module
    'C': 25,  # Sous-Module
    'D': 50,  # Tâche
    'E': 12,  # Type
    'F': 15,  # Composant
    'G': 45,  # Fichier
    'H': 15,  # Dépendances
    'I': 12,  # Durée
    'J': 12,  # Priorité
    'K': 12,  # Statut
    'L': 60   # Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 6. Formater les lignes de données
for row_idx in range(2, ws.max_row + 1):
    # Récupérer les valeurs
    sous_module = ws[f'C{row_idx}'].value
    priorite = ws[f'J{row_idx}'].value
    
    # Couleur de fond selon la phase
    bg_color = PHASE_COLORS.get(sous_module, "FFFFFF")
    
    # Appliquer le style à toute la ligne
    for col_idx in range(1, 13):  # Colonnes A à L
        cell = ws.cell(row=row_idx, column=col_idx)
        
        # Couleur de fond
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Bordure
        cell.border = thin_border
        
        # Alignement
        if col_idx == 1:  # ID - centré
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [5, 10, 11]:  # Type, Priorité, Statut - centré
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 9:  # Durée - centré
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Colorer la priorité
    priorite_cell = ws[f'J{row_idx}']
    if priorite == 'HAUTE':
        priorite_cell.font = Font(bold=True, color="FFFFFF")
        priorite_cell.fill = PatternFill(start_color=HAUTE_COLOR, end_color=HAUTE_COLOR, fill_type="solid")
    elif priorite == 'MOYENNE':
        priorite_cell.font = Font(bold=True, color="FFFFFF")
        priorite_cell.fill = PatternFill(start_color=MOYENNE_COLOR, end_color=MOYENNE_COLOR, fill_type="solid")
    elif priorite == 'BASSE':
        priorite_cell.font = Font(bold=True, color="FFFFFF")
        priorite_cell.fill = PatternFill(start_color=BASSE_COLOR, end_color=BASSE_COLOR, fill_type="solid")
    
    # ID en gras
    ws[f'A{row_idx}'].font = Font(bold=True)

# 7. Figer la première ligne
ws.freeze_panes = 'A2'

# 8. Ajouter des filtres automatiques
ws.auto_filter.ref = ws.dimensions

# 9. Ajuster la hauteur des lignes
ws.row_dimensions[1].height = 30  # En-tête
for row_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 35  # Lignes de données

# 10. Créer une feuille "Résumé Phases"
ws_phases = wb.create_sheet("Résumé Phases", 0)

# Données des phases
phases_data = [
    ["Phase", "Nom", "Tâches", "Durée (h)", "Priorité", "Statut"],
    [1, "Base de Données", "1-10", 4, "HAUTE", "À faire"],
    [2, "Services Backend", "11-14", 6, "HAUTE", "À faire"],
    [3, "API Routes", "15-18", 4.5, "HAUTE", "À faire"],
    [4, "Frontend - Scanner", "19-23", 3.5, "HAUTE", "À faire"],
    [5, "Frontend - Gestion", "24-27", 3.8, "MOYENNE", "À faire"],
    [6, "Dashboard Stats", "28-30", 3, "MOYENNE", "À faire"],
    [7, "Tests & Validation", "31-41", 5.25, "HAUTE", "À faire"],
    [8, "Documentation", "42-45", 1.75, "MOYENNE", "À faire"],
    ["", "TOTAL", "", 31.8, "", ""]
]

# Écrire les données
for row_idx, row_data in enumerate(phases_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_phases.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        
        # Formater l'en-tête
        if row_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formater la ligne TOTAL
        elif row_idx == len(phases_data):
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Lignes normales
        else:
            if col_idx in [1, 3, 4]:  # Phase, Tâches, Durée - centré
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

# Largeur colonnes feuille phases
ws_phases.column_dimensions['A'].width = 10
ws_phases.column_dimensions['B'].width = 25
ws_phases.column_dimensions['C'].width = 12
ws_phases.column_dimensions['D'].width = 12
ws_phases.column_dimensions['E'].width = 12
ws_phases.column_dimensions['F'].width = 15

# 11. Créer une feuille "État Modules"
ws_modules = wb.create_sheet("État Modules")

modules_data = [
    ["Module", "Backend", "Frontend", "Progression (%)", "Statut", "Ce qui manque"],
    ["Authentification", "✓", "✓", 100, "✅ COMPLET", "Rien"],
    ["Formations", "✓", "✓", 100, "✅ COMPLET", "Rien"],
    ["Salles", "✓", "✓", 100, "✅ COMPLET", "Rien"],
    ["Créneaux Horaires", "✓", "✓", 100, "✅ COMPLET", "Rien"],
    ["Formateurs", "✓", "✓", 100, "✅ COMPLET", "Rien"],
    ["Sessions", "✓", "✓", 100, "✅ COMPLET", "Rien"],
    ["Inscriptions", "✓", "✓", 90, "🟡 PARTIEL", "Génération badge QR backend"],
    ["Étudiants", "✓", "✓", 85, "🟡 PARTIEL", "Champs QR backend"],
    ["Affectations", "✓", "✓", 85, "🟡 PARTIEL", "Page frontend à nettoyer"],
    ["Échéanciers", "✓", "✓", 95, "🟡 PARTIEL", "Auto-génération"],
    ["Paiements", "✓", "✓", 80, "🟡 PARTIEL", "Nettoyage doublons"],
    ["Retards Paiement", "✓", "✓", 90, "🟡 PARTIEL", "Tests complets"],
    ["Dashboard", "✓", "✓", 70, "🟡 PARTIEL", "Widgets présences"],
    ["Système QR Code", "❌", "❌", 0, "❌ À CRÉER", "Tout à créer"],
    ["Présences", "❌", "❌", 0, "❌ À CRÉER", "Tout à créer"],
    ["Scan QR", "❌", "❌", 0, "❌ À CRÉER", "Tout à créer"],
    ["Rapports Présences", "❌", "❌", 0, "❌ À CRÉER", "Tout à créer"]
]

for row_idx, row_data in enumerate(modules_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_modules.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        
        if row_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            # Colorer selon progression
            if row_data[3] == 100:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif row_data[3] >= 80:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif row_data[3] > 0:
                cell.fill = PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            if col_idx in [2, 3, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

ws_modules.column_dimensions['A'].width = 25
ws_modules.column_dimensions['B'].width = 12
ws_modules.column_dimensions['C'].width = 12
ws_modules.column_dimensions['D'].width = 15
ws_modules.column_dimensions['E'].width = 15
ws_modules.column_dimensions['F'].width = 35

# 12. Sauvegarder
print("💾 Sauvegarde du fichier...")
wb.save('PROJECT_PLAN.xlsx')

print("✅ Fichier PROJECT_PLAN.xlsx créé avec succès!")
print("📊 Feuilles créées:")
print("   - Résumé Phases (Vue d'ensemble)")
print("   - État Modules (Progression actuelle)")
print("   - Plan Projet (45 tâches détaillées)")
print("\n🎯 Vous pouvez maintenant ouvrir PROJECT_PLAN.xlsx dans Excel!")
